import json
import re
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"Z:\Wagmi\Wagmi Files\zich scripts(AutoRecovered).xlsx")
GOOGLE_SOURCE = Path(r"C:\Users\jabdu\Documents\Codex\2026-05-27\files-mentioned-by-the-user-zich\google-sheet-export.xlsx")
OUT = Path(r"C:\Users\jabdu\Documents\Codex\2026-05-27\files-mentioned-by-the-user-zich\data.js")


def clean(value):
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    return text.strip()


def is_heading(text):
    if not text:
        return False
    if len(text) > 90:
        return False
    letters = re.sub(r"[^A-Za-z]+", "", text)
    if len(letters) < 3:
        return False
    uppercase_ratio = sum(1 for ch in letters if ch.isupper()) / len(letters)
    return uppercase_ratio > 0.72


def title_from_text(text):
    first = re.split(r"[\n.!?]", text.strip(), maxsplit=1)[0]
    first = re.sub(r"\s+", " ", first).strip()
    if len(first) > 72:
        first = first[:69].rstrip() + "..."
    return first or "Untitled"


wb = load_workbook(SOURCE, data_only=True, read_only=True)
records = []
inventory = []
specs = []
links = []
sheet_meta = []
products = []
tickets = []
daily_tracker = []
replacement_codes = []
formulas = []

for ws in wb.worksheets:
    sheet_meta.append({"name": ws.title, "rows": ws.max_row, "columns": ws.max_column})
    column_heading = {}
    current_section = "General"
    header_values = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    if ws.title == "STOCKS UPDATE":
        current_sku = ""
        for row in range(2, ws.max_row + 1):
            sku = clean(ws.cell(row, 1).value) or current_sku
            current_sku = sku or current_sku
            warehouse = clean(ws.cell(row, 2).value)
            qty = ws.cell(row, 3).value
            price = ws.cell(row, 4).value
            if sku or warehouse or qty is not None or price is not None:
                inventory.append(
                    {
                        "sku": sku,
                        "warehouse": warehouse,
                        "quantity": qty if isinstance(qty, (int, float)) else clean(qty),
                        "previousPrice": price if isinstance(price, (int, float)) else clean(price),
                    }
                )
        continue

    if ws.title == "noise level":
        models = [clean(ws.cell(1, c).value) for c in range(2, ws.max_column + 1)]
        for row in range(2, ws.max_row + 1):
            metric = clean(ws.cell(row, 1).value)
            if not metric:
                continue
            values = {}
            for idx, model in enumerate(models, start=2):
                if not model or model.startswith("Unnamed"):
                    continue
                value = clean(ws.cell(row, idx).value)
                if value:
                    values[model] = value
            if values:
                specs.append({"metric": metric, "values": values})
        continue

    for row in range(1, ws.max_row + 1):
        row_values = [clean(ws.cell(row, c).value) for c in range(1, ws.max_column + 1)]
        nonempty = [v for v in row_values if v]
        if len(nonempty) == 1 and is_heading(nonempty[0]):
            current_section = nonempty[0]

        for col, text in enumerate(row_values, start=1):
            if not text:
                continue

            header = header_values[col - 1] if col - 1 < len(header_values) else ""
            if is_heading(text):
                column_heading[col] = text
                if len(nonempty) <= 3:
                    continue

            category = column_heading.get(col) or (header if header and not header.startswith("Unnamed") else current_section)
            if header and header.startswith("Unnamed"):
                header = ""

            url_match = re.search(r"https?://\S+", text)
            if url_match:
                links.append(
                    {
                        "sheet": ws.title,
                        "label": title_from_text(text.replace(url_match.group(0), "")) or url_match.group(0),
                        "url": url_match.group(0).rstrip(").,"),
                        "context": text,
                    }
                )

            if len(text) < 18 and is_heading(text):
                continue

            records.append(
                {
                    "id": f"{ws.title}-{row}-{col}",
                    "sheet": ws.title,
                    "section": category or "General",
                    "title": title_from_text(text),
                    "body": text,
                    "cell": f"{ws.cell(row, col).coordinate}",
                }
            )

if GOOGLE_SOURCE.exists():
    g_values = load_workbook(GOOGLE_SOURCE, data_only=True, read_only=False)
    g_formulas = load_workbook(GOOGLE_SOURCE, data_only=False, read_only=False)
    for ws_formula in g_formulas.worksheets:
        ws_value = g_values[ws_formula.title]
        sheet_meta.append({"name": f"Google - {ws_formula.title}", "rows": ws_formula.max_row or 0, "columns": ws_formula.max_column or 0})
        for row in range(1, (ws_formula.max_row or 0) + 1):
            for col in range(1, (ws_formula.max_column or 0) + 1):
                formula = ws_formula.cell(row, col).value
                if isinstance(formula, str) and formula.startswith("="):
                    formulas.append(
                        {
                            "sheet": ws_formula.title,
                            "cell": ws_formula.cell(row, col).coordinate,
                            "formula": formula,
                            "value": clean(ws_value.cell(row, col).value),
                        }
                    )

    if "Canned Responses Clean" in g_values.sheetnames:
        ws = g_values["Canned Responses Clean"]
        for row in range(2, (ws.max_row or 0) + 1):
            category = clean(ws.cell(row, 1).value) or "Google Sheet"
            title = clean(ws.cell(row, 2).value)
            body = clean(ws.cell(row, 3).value)
            notes = clean(ws.cell(row, 4).value)
            if title or body:
                text = body or title
                records.append(
                    {
                        "id": f"Google-Canned-{row}",
                        "sheet": "Google - Canned Responses Clean",
                        "section": category,
                        "title": title or title_from_text(text),
                        "body": text + (f"\n\nNotes: {notes}" if notes else ""),
                        "cell": f"A{row}:D{row}",
                    }
                )

    if "Ticket Library Clean" in g_values.sheetnames:
        ws = g_values["Ticket Library Clean"]
        for row in range(2, (ws.max_row or 0) + 1):
            link = clean(ws.cell(row, 1).value)
            ticket_id = clean(ws.cell(row, 2).value)
            ticket_type = clean(ws.cell(row, 3).value)
            sku = clean(ws.cell(row, 4).value)
            category = clean(ws.cell(row, 5).value) or "Ticket Library"
            notes = clean(ws.cell(row, 6).value)
            if link or ticket_type or ticket_id:
                ticket = {
                    "link": link,
                    "ticketId": ticket_id,
                    "type": ticket_type,
                    "sku": sku,
                    "category": category,
                    "notes": notes,
                }
                tickets.append(ticket)
                records.append(
                    {
                        "id": f"Google-Ticket-{row}",
                        "sheet": "Google - Ticket Library Clean",
                        "section": category,
                        "title": ticket_type or f"Ticket {ticket_id}",
                        "body": f"Ticket ID: {ticket_id}\nSKU / Model: {sku}\nLink: {link}" + (f"\nNotes: {notes}" if notes else ""),
                        "cell": f"A{row}:F{row}",
                    }
                )
                if link:
                    links.append({"sheet": "Google - Ticket Library Clean", "label": ticket_type or f"Ticket {ticket_id}", "url": link, "context": category})

    if "PRODUCT DIMENSIONS" in g_values.sheetnames:
        ws = g_values["PRODUCT DIMENSIONS"]
        headers = [clean(ws.cell(1, col).value) or f"Column {col}" for col in range(1, (ws.max_column or 0) + 1)]
        for row in range(2, (ws.max_row or 0) + 1):
            values = {headers[col - 1]: clean(ws.cell(row, col).value) for col in range(1, (ws.max_column or 0) + 1)}
            if any(values.values()):
                products.append(values)

    for code_sheet in ["Replacement Codes", "Copy of Replacement Category an"]:
        if code_sheet in g_values.sheetnames:
            ws_v = g_values[code_sheet]
            ws_f = g_formulas[code_sheet]
            for row in range(2, (ws_v.max_row or 0) + 1):
                category = clean(ws_v.cell(row, 1).value)
                reason = clean(ws_v.cell(row, 2).value)
                description = clean(ws_v.cell(row, 3).value)
                order_number = clean(ws_v.cell(row, 4).value)
                generated = clean(ws_v.cell(row, 5).value)
                formula = clean(ws_f.cell(row, 5).value)
                if category or reason or description:
                    replacement_codes.append(
                        {
                            "sheet": code_sheet,
                            "category": category,
                            "reasonCode": reason,
                            "description": description,
                            "orderNumber": order_number,
                            "generatedCode": generated,
                            "formula": formula,
                        }
                    )

    if "Daily Tracker Clean" in g_values.sheetnames:
        ws = g_values["Daily Tracker Clean"]
        for row in range(2, (ws.max_row or 0) + 1):
            entry = {
                "date": clean(ws.cell(row, 1).value),
                "link": clean(ws.cell(row, 2).value),
                "ticketId": clean(ws.cell(row, 3).value),
                "status": clean(ws.cell(row, 4).value),
                "notes": clean(ws.cell(row, 5).value),
            }
            if any(entry.values()):
                daily_tracker.append(entry)

payload = {
    "source": SOURCE.name,
    "generatedFrom": str(SOURCE),
    "sheets": sheet_meta,
    "records": records,
    "inventory": inventory,
    "specs": specs,
    "links": links,
    "products": products,
    "tickets": tickets,
    "dailyTracker": daily_tracker,
    "replacementCodes": replacement_codes,
    "formulas": formulas,
}

OUT.write_text("window.KB_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")
print(
    json.dumps(
        {
            "records": len(records),
            "inventory": len(inventory),
            "specRows": len(specs),
            "links": len(links),
            "products": len(products),
            "tickets": len(tickets),
            "replacementCodes": len(replacement_codes),
            "formulas": len(formulas),
            "sheets": [s["name"] for s in sheet_meta],
        },
        indent=2,
    )
)
