import json
from pathlib import Path

from openpyxl import load_workbook

path = Path(r"C:\Users\jabdu\Documents\Codex\2026-05-27\files-mentioned-by-the-user-zich\google-sheet-export.xlsx")
wb_values = load_workbook(path, data_only=True, read_only=False)
wb_formulas = load_workbook(path, data_only=False, read_only=False)

summary = []
for ws_formula in wb_formulas.worksheets:
    ws_value = wb_values[ws_formula.title]
    total_rows = ws_formula.max_row or 0
    total_cols = ws_formula.max_column or 0
    formula_count = 0
    samples = []
    rows = []
    max_rows = min(total_rows, 8)
    max_cols = min(total_cols, 8)
    for row in range(1, max_rows + 1):
        out = []
        for col in range(1, max_cols + 1):
            cell_formula = ws_formula.cell(row, col).value
            cell_value = ws_value.cell(row, col).value
            if isinstance(cell_formula, str) and cell_formula.startswith("="):
                formula_count += 1
                if len(samples) < 8:
                    samples.append({"cell": ws_formula.cell(row, col).coordinate, "formula": cell_formula, "value": cell_value})
            out.append(cell_value if cell_value is not None else cell_formula)
        rows.append(out)
    for row in range(max_rows + 1, total_rows + 1):
        for col in range(1, total_cols + 1):
            cell_formula = ws_formula.cell(row, col).value
            if isinstance(cell_formula, str) and cell_formula.startswith("="):
                formula_count += 1
                if len(samples) < 8:
                    samples.append({"cell": ws_formula.cell(row, col).coordinate, "formula": cell_formula, "value": ws_value.cell(row, col).value})
    summary.append(
        {
            "sheet": ws_formula.title,
            "rows": total_rows,
            "columns": total_cols,
            "formulas": formula_count,
            "formulaSamples": samples,
            "preview": rows,
        }
    )

print(json.dumps(summary, ensure_ascii=True, indent=2, default=str))
